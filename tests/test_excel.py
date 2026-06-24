"""
Tests for the Excel Report Generation Service.

Tests report creation with mock data and verifies structure.
"""

import os
import sys
from datetime import datetime

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.services.excel_service import ExcelService


@pytest.fixture
def excel_service():
    """Create an ExcelService instance."""
    return ExcelService()


@pytest.fixture
def mock_results():
    """Create mock scan results."""
    return [
        {
            "symbol": "TESTCO",
            "company_name": "Test Company Ltd",
            "ipo_year": 2020,
            "ipo_first_month_high": 150.0,
            "breakout_month": "2020-06",
            "breakout_close": 180.0,
            "current_price": 350.0,
            "pct_above_ipo_high": 133.33,
            "listing_date": datetime(2020, 3, 15),
            "qualified": True,
        },
        {
            "symbol": "SAMPLE",
            "company_name": "Sample Corp",
            "ipo_year": 2020,
            "ipo_first_month_high": 200.0,
            "breakout_month": "2021-01",
            "breakout_close": 250.0,
            "current_price": 500.0,
            "pct_above_ipo_high": 150.0,
            "listing_date": datetime(2020, 7, 1),
            "qualified": True,
        },
        {
            "symbol": "NOBREAK",
            "company_name": "No Breakout Inc",
            "ipo_year": 2020,
            "ipo_first_month_high": 300.0,
            "breakout_month": None,
            "breakout_close": None,
            "current_price": 250.0,
            "pct_above_ipo_high": None,
            "listing_date": datetime(2020, 9, 10),
            "qualified": False,
        },
    ]


@pytest.fixture
def mock_qualified(mock_results):
    """Get only qualified results."""
    return [r for r in mock_results if r["qualified"]]


class TestExcelGeneration:
    """Test Excel report generation."""

    def test_report_creates_file(self, excel_service, mock_results, mock_qualified):
        """Test that generate_report creates an Excel file."""
        filepath = excel_service.generate_report(
            year=2020,
            scan_id="test-scan-001",
            results=mock_results,
            qualified_results=mock_qualified,
        )

        assert os.path.exists(filepath)
        assert filepath.endswith(".xlsx")
        assert "IPO_Breakout_2020" in filepath

        # Clean up
        os.remove(filepath)

    def test_report_has_three_sheets(self, excel_service, mock_results, mock_qualified):
        """Test that the report has all three sheets."""
        filepath = excel_service.generate_report(
            year=2020,
            scan_id="test-scan-002",
            results=mock_results,
            qualified_results=mock_qualified,
        )

        wb = load_workbook(filepath)
        sheet_names = wb.sheetnames

        assert "Qualified Stocks" in sheet_names
        assert "All Stocks" in sheet_names
        assert "Summary" in sheet_names

        wb.close()
        os.remove(filepath)

    def test_qualified_sheet_headers(self, excel_service, mock_results, mock_qualified):
        """Test that the qualified sheet has correct headers."""
        filepath = excel_service.generate_report(
            year=2020,
            scan_id="test-scan-003",
            results=mock_results,
            qualified_results=mock_qualified,
        )

        wb = load_workbook(filepath)
        ws = wb["Qualified Stocks"]

        expected_headers = [
            "Symbol", "Company Name", "IPO Year", "IPO First Month High",
            "Breakout Month", "Breakout Close", "Previous Month Close", "Current Price",
            "% Above IPO High", "Listing Date",
        ]

        # Headers are on row 4 (after title and subtitle)
        for col_idx, expected in enumerate(expected_headers, 1):
            cell_value = ws.cell(row=4, column=col_idx).value
            assert cell_value == expected, f"Column {col_idx}: expected '{expected}', got '{cell_value}'"

        wb.close()
        os.remove(filepath)

    def test_qualified_sheet_data(self, excel_service, mock_results, mock_qualified):
        """Test that data rows are correctly populated."""
        filepath = excel_service.generate_report(
            year=2020,
            scan_id="test-scan-004",
            results=mock_results,
            qualified_results=mock_qualified,
        )

        wb = load_workbook(filepath)
        ws = wb["Qualified Stocks"]

        # Data starts at row 5
        # Should have 2 qualified stocks
        data_rows = []
        for row in range(5, ws.max_row + 1):
            symbol = ws.cell(row=row, column=1).value
            if symbol:
                data_rows.append(symbol)

        assert len(data_rows) == 2
        assert set(data_rows) == {"TESTCO", "SAMPLE"} or set(data_rows) == {"SAMPLE", "TESTCO"}

        wb.close()
        os.remove(filepath)

    def test_summary_sheet_content(self, excel_service, mock_results, mock_qualified):
        """Test that the summary sheet has correct statistics."""
        filepath = excel_service.generate_report(
            year=2020,
            scan_id="test-scan-005",
            results=mock_results,
            qualified_results=mock_qualified,
        )

        wb = load_workbook(filepath)
        ws = wb["Summary"]

        # Check title contains year
        title = ws["A1"].value
        assert "2020" in str(title)

        # Check stats area (row 4 onwards)
        found_total = False
        found_qualified = False
        for row in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=row, column=1).value
            cell_b = ws.cell(row=row, column=2).value
            if cell_a == "Total Stocks Scanned":
                assert cell_b == 3
                found_total = True
            elif cell_a == "Qualified Stocks":
                assert cell_b == 2
                found_qualified = True

        assert found_total, "Total Stocks Scanned not found in summary"
        assert found_qualified, "Qualified Stocks count not found in summary"

        wb.close()
        os.remove(filepath)

    def test_empty_results(self, excel_service):
        """Test report generation with empty results."""
        filepath = excel_service.generate_report(
            year=2020,
            scan_id="test-scan-006",
            results=[],
            qualified_results=[],
        )

        assert os.path.exists(filepath)

        wb = load_workbook(filepath)
        assert len(wb.sheetnames) == 3

        wb.close()
        os.remove(filepath)

    def test_all_stocks_sheet(self, excel_service, mock_results, mock_qualified):
        """Test that the All Stocks sheet includes all results."""
        filepath = excel_service.generate_report(
            year=2020,
            scan_id="test-scan-007",
            results=mock_results,
            qualified_results=mock_qualified,
        )

        wb = load_workbook(filepath)
        ws = wb["All Stocks"]

        # Count data rows (skip header)
        data_rows = 0
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value:
                data_rows += 1

        assert data_rows == 3  # All 3 results (2 qualified + 1 unqualified)

        wb.close()
        os.remove(filepath)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
