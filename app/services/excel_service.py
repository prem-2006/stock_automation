"""
Excel Report Generation Service.

Creates professionally formatted Excel reports with:
- Detailed stock screening results
- Summary statistics sheet
- Conditional formatting and styling
"""

import os
from datetime import datetime
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

from app.utils.logger import get_logger

logger = get_logger("excel_service")

# Report output directory — use /tmp on Vercel (only writable dir in serverless)
IS_VERCEL = os.environ.get("VERCEL", "") == "1" or os.environ.get("VERCEL_ENV") is not None
REPORTS_DIR = "/tmp/reports" if IS_VERCEL else os.path.join("app", "static", "reports")


class ExcelService:
    """Service for generating styled Excel reports."""

    def __init__(self):
        os.makedirs(REPORTS_DIR, exist_ok=True)

    def generate_report(
        self,
        year: int,
        scan_id: str,
        results: List[Dict],
        qualified_results: List[Dict],
    ) -> str:
        """
        Generate a comprehensive Excel report for a stock scan.

        Args:
            year: IPO year scanned
            scan_id: Scan job ID
            results: All stock results (qualified + unqualified)
            qualified_results: Only qualified stocks

        Returns:
            File path of the generated Excel report
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"IPO_Breakout_{year}_{timestamp}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)

        wb = Workbook()

        # Sheet 1: Qualified Results
        self._create_results_sheet(wb, qualified_results, year)

        # Sheet 2: All Results
        self._create_all_results_sheet(wb, results, year)

        # Sheet 3: Summary Statistics
        self._create_summary_sheet(wb, results, qualified_results, year)

        wb.save(filepath)
        logger.info(f"Excel report generated: {filepath}")

        return filepath

    def _create_results_sheet(self, wb: Workbook, results: List[Dict], year: int):
        """Create the main qualified results sheet with professional styling."""
        ws = wb.active
        ws.title = "Qualified Stocks"

        # Define styles
        header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        title_font = Font(name="Calibri", bold=True, size=16, color="1B4F72")
        subtitle_font = Font(name="Calibri", size=11, color="555555", italic=True)

        data_font = Font(name="Calibri", size=11)
        data_alignment = Alignment(horizontal="center", vertical="center")

        green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        alt_row_fill = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="D5D8DC"),
            right=Side(style="thin", color="D5D8DC"),
            top=Side(style="thin", color="D5D8DC"),
            bottom=Side(style="thin", color="D5D8DC"),
        )

        # Title
        ws.merge_cells("A1:I1")
        ws["A1"] = f"📊 IPO Breakout Scan Report — Year {year}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        # Subtitle
        ws.merge_cells("A2:I2")
        ws["A2"] = f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}  |  {len(results)} Qualified Stocks"
        ws["A2"].font = subtitle_font
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 25

        # Empty row
        ws.row_dimensions[3].height = 10

        # Headers
        headers = [
            "Symbol",
            "Company Name",
            "IPO Year",
            "IPO First Month High",
            "Breakout Month",
            "Breakout Close",
            "Previous Month Close",
            "Current Price",
            "% Above IPO High",
            "Listing Date",
        ]

        header_row = 4
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[header_row].height = 35

        # Sort by % above IPO high descending
        sorted_results = sorted(
            results,
            key=lambda x: x.get("pct_above_ipo_high", 0) or 0,
            reverse=True,
        )

        # Data rows
        for row_idx, result in enumerate(sorted_results, header_row + 1):
            listing_date = result.get("listing_date")
            if hasattr(listing_date, "strftime"):
                listing_date_str = listing_date.strftime("%Y-%m-%d")
            elif listing_date:
                listing_date_str = str(listing_date)[:10]
            else:
                listing_date_str = "N/A"

            row_data = [
                result.get("symbol", ""),
                result.get("company_name", ""),
                result.get("ipo_year", year),
                result.get("ipo_first_month_high"),
                result.get("breakout_month", "N/A"),
                result.get("breakout_close"),
                result.get("previous_month_close"),
                result.get("current_price"),
                result.get("pct_above_ipo_high"),
                listing_date_str,
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

                # Alternate row coloring
                if (row_idx - header_row) % 2 == 0:
                    cell.fill = alt_row_fill

            # Format percentage column
            pct_cell = ws.cell(row=row_idx, column=9)
            if pct_cell.value and isinstance(pct_cell.value, (int, float)):
                pct_cell.number_format = "0.00%"
                pct_cell.value = pct_cell.value / 100  # Convert to decimal for Excel %

                # Green highlight for high performers
                if pct_cell.value > 1.0:  # > 100%
                    pct_cell.fill = green_fill

            # Format currency columns
            for col in [4, 6, 7, 8]:  # IPO High, Breakout Close, Prev Month Close, Current Price
                cell = ws.cell(row=row_idx, column=col)
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = "₹#,##0.00"

        # Auto-fit column widths
        column_widths = [15, 35, 12, 22, 18, 18, 22, 16, 20, 15]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # Freeze header row
        ws.freeze_panes = "A5"

    def _create_all_results_sheet(self, wb: Workbook, results: List[Dict], year: int):
        """Create a sheet with all scanned stocks (qualified and unqualified)."""
        ws = wb.create_sheet(title="All Stocks")

        # Styles
        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        qualified_fill = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
        not_qualified_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="D5D8DC"),
            right=Side(style="thin", color="D5D8DC"),
            top=Side(style="thin", color="D5D8DC"),
            bottom=Side(style="thin", color="D5D8DC"),
        )

        headers = [
            "Symbol", "Company Name", "Qualified", "IPO First Month High",
            "Breakout Month", "Breakout Close", "Previous Month Close", "Current Price", "% Above IPO High",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, result in enumerate(results, 2):
            row_data = [
                result.get("symbol", ""),
                result.get("company_name", ""),
                "Yes" if result.get("qualified") else "No",
                result.get("ipo_first_month_high"),
                result.get("breakout_month", "N/A"),
                result.get("breakout_close"),
                result.get("previous_month_close"),
                result.get("current_price"),
                result.get("pct_above_ipo_high"),
            ]

            fill = qualified_fill if result.get("qualified") else not_qualified_fill

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        # Auto-fit
        column_widths = [15, 35, 12, 22, 18, 18, 22, 16, 20]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        ws.freeze_panes = "A2"

    def _create_summary_sheet(
        self, wb: Workbook, all_results: List[Dict], qualified: List[Dict], year: int
    ):
        """Create a summary statistics sheet."""
        ws = wb.create_sheet(title="Summary")

        # Styles
        title_font = Font(name="Calibri", bold=True, size=16, color="1B4F72")
        section_font = Font(name="Calibri", bold=True, size=13, color="2C3E50")
        label_font = Font(name="Calibri", size=12, color="555555")
        value_font = Font(name="Calibri", bold=True, size=14, color="1B4F72")
        accent_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = f"📈 IPO Breakout Scan Summary — {year}"
        ws["A1"].font = title_font
        ws.row_dimensions[1].height = 40

        # Summary Statistics
        total = len(all_results)
        qual_count = len(qualified)
        qual_pct = (qual_count / total * 100) if total > 0 else 0

        stats = [
            ("Total Stocks Scanned", total),
            ("Qualified Stocks", qual_count),
            ("Qualification Percentage", f"{qual_pct:.1f}%"),
            ("Scan Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ]

        row = 3
        ws.cell(row=row, column=1, value="📊 Overall Statistics").font = section_font
        row += 1

        for label, value in stats:
            ws.cell(row=row, column=1, value=label).font = label_font
            val_cell = ws.cell(row=row, column=2, value=value)
            val_cell.font = value_font
            val_cell.fill = accent_fill
            row += 1

        # Top 10 Strongest Stocks
        row += 1
        ws.cell(row=row, column=1, value="🏆 Top 10 Strongest Stocks").font = section_font
        row += 1

        top_10_headers = ["Rank", "Symbol", "Company", "% Above IPO High"]
        header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
        header_font_white = Font(name="Calibri", bold=True, size=11, color="FFFFFF")

        for col_idx, header in enumerate(top_10_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row += 1
        sorted_qualified = sorted(
            qualified,
            key=lambda x: x.get("pct_above_ipo_high", 0) or 0,
            reverse=True,
        )

        for rank, stock in enumerate(sorted_qualified[:10], 1):
            ws.cell(row=row, column=1, value=rank).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=stock.get("symbol", ""))
            ws.cell(row=row, column=3, value=stock.get("company_name", ""))
            pct = stock.get("pct_above_ipo_high", 0) or 0
            pct_cell = ws.cell(row=row, column=4, value=f"{pct:.2f}%")
            pct_cell.alignment = Alignment(horizontal="center")

            # Gold/Silver/Bronze for top 3
            if rank <= 3:
                medal_colors = {1: "FFD700", 2: "C0C0C0", 3: "CD7F32"}
                medal_fill = PatternFill(
                    start_color=medal_colors[rank],
                    end_color=medal_colors[rank],
                    fill_type="solid",
                )
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = medal_fill

            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 20
